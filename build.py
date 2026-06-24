#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate index.html (desktop), mobile.html (phone), and rooms.xlsx from a
single data source. Rooms near Kyung Hee University (Seoul) for a one-semester
exchange student. Snapshot date: 2026-06-23. Prices on platforms change; treat
as a guide."""

import html
import json
import os
from urllib.parse import quote

SNAPSHOT = "2026-06-23"

def naver(q):
    return "https://map.naver.com/p/search/" + quote(q)

def ziptoss_en(area):
    return "https://ziptoss.com/en/map/" + quote(area) + "?contract=monthly"

# ---------------------------------------------------------------- DATA
# Fallback listings (used when listings.json is absent or unreadable).
FALLBACK_FEATURED = [
    {
        "name": "Hoegi-dong full-option studio (KHU 2 min)",
        "platform": "33m2 (English app)",
        "type": "Furnished studio",
        "rent": "₩740,000 / mo  (₩170,000 / week)",
        "deposit": "₩330,000 (fixed)",
        "size": "13 m² (4 pyeong)",
        "station": "Hoegi Stn (Line 1 / Gyeongui-Jungang / Gyeongchun)",
        "commute": "KHU back gate 2-min walk",
        "options": "Bed, Wardrobe, Desk, Fridge, Microwave, Washer, A/C, Smart lock, CCTV; electricity+water+gas+internet ALL included",
        "english": "Yes — English app, foreign card OK",
        "link": "https://web.33m2.co.kr/guest/room/47052",
        "naver": naver("경희대학교 후문"),
        "notes": "Cheapest turnkey option right at the campus back gate. Min 2 weeks; 5% off for 8+ weeks. No ARC needed.",
    },
    {
        "name": "Hoegi-dong full-option studio (KHU 2 min, larger)",
        "platform": "33m2 (English app)",
        "type": "Furnished studio",
        "rent": "₩880,000 / mo  (₩220,000 / week)",
        "deposit": "₩330,000 (fixed)",
        "size": "20 m² (6 pyeong)",
        "station": "Hoegi Stn (Line 1)",
        "commute": "KHU back gate 2-min walk",
        "options": "Bed, Wardrobe, Desk, Fridge, Microwave, Gas range, Washer, Water purifier, A/C, WiFi; utilities ALL included",
        "english": "Yes — English app, foreign card OK",
        "link": "https://web.33m2.co.kr/guest/room/47054",
        "naver": naver("경희대학교 후문"),
        "notes": "More space, still walk-to-class. Min 2 weeks; 5% off for 8+ weeks.",
    },
    {
        "name": "Hoegi / KHU open studio (deposit-based)",
        "platform": "Daangn (당근, KR only)",
        "type": "Open studio",
        "rent": "₩500,000 / mo",
        "deposit": "₩5,000,000",
        "size": "Studio",
        "station": "Hoegi Stn (Line 1)",
        "commute": "~10–15 min to KHU",
        "options": "Full-option typical for the building",
        "english": "No — needs Korean or an agent; bring a helper",
        "link": "https://www.daangn.com/kr/realty/?in=%ED%9A%8C%EA%B8%B0%EB%8F%99-97&salesType=one_room",
        "naver": naver("회기역"),
        "notes": "Lowest monthly rent IF the ₩5M deposit is affordable. Standard 1-yr leases; ask about a one-semester (short) contract.",
    },
]

# ── Translation tables ────────────────────────────────────────────────────────
_TYPE_MAP = {
    "일반주택": "House",
    "오피스텔": "Officetel",
    "기타": "Other",
    "상가주택": "Commercial house",
}
_DONG_MAP = {
    "삼선동5가": "Samseon-dong 5-ga",
    "휘경동": "Hwigyeong-dong",
    "종암동": "Jongam-dong",
    "이문동": "Imun-dong",
    "하월곡동": "Hawolgok-dong",
    "회기동": "Hoegi-dong",
    "용두동": "Yongdu-dong",
    "청량리동": "Cheongnyangni-dong",
    "전농동": "Jeonnong-dong",
    "제기동": "Jegi-dong",
    "상월곡동": "Sangwolgok-dong",
    "월계동": "Wolgye-dong",
    "안암동5가": "Anam-dong 5-ga",
    "안암동2가": "Anam-dong 2-ga",
    "석관동": "Seokgwan-dong",
    "보문동5가": "Bomun-dong 5-ga",
    "보문동6가": "Bomun-dong 6-ga",
}
# Ordered longest-first to prevent partial matches.
_STATION_MAP = [
    ("회기역 (경춘선)", "Hoegi Stn (Gyeongchun)"),
    ("청량리역 (경의중앙선)", "Cheongnyangni Stn (Gyeongui-Jungang)"),
    ("외대앞역, 회기역", "HUFS Stn, Hoegi Stn"),
    ("회기역, 외대앞역", "Hoegi Stn, HUFS Stn"),
    ("외대앞역", "HUFS Stn"),
    ("회기역", "Hoegi Stn"),
    ("상월곡역", "Sangwolgok Stn"),
    ("월곡역", "Wolgok Stn"),
    ("제기동역", "Jegi-dong Stn"),
    ("청량리역", "Cheongnyangni Stn"),
    ("고려대역", "Korea Univ. Stn"),
    ("안암역", "Anam Stn"),
    ("신이문역", "Sinimun Stn"),
    ("창신역", "Changshin Stn"),
    ("월계역", "Wolgye Stn"),
    ("보문역", "Bomun Stn"),
]
_OPTIONS_MAP = {
    "에어컨": "A/C",
    "세탁기": "Washer",
    "냉장고": "Fridge",
    "인덕션": "Induction",
    "가스레인지": "Gas range",
    "전자도어락": "Smart lock",
    "침대": "Bed",
    "책상": "Desk",
    "옷장": "Wardrobe",
    "신발장": "Shoe rack",
    "의자": "Chair",
    "전자레인지": "Microwave",
    "베란다": "Balcony",
    "다용도실": "Utility room",
}
_M2_TITLE_MAP = {
    "특가 🛎️ 외대2분 풀옵": "Deal: 2 min to HUFS, full-option",
    "외대역도보3분 경희대": "3 min walk to HUFS Stn · near KHU",
    "회기삼육 조용한 독채": "Quiet detached unit · Hoegi",
    "아늑한방!!회기역8분": "Cozy room · 8 min to Hoegi Stn",
    "편안한방😋경희대3분!": "Comfortable room · 3 min to KHU",
    "🎈회기역 도보3분 넓음": "Spacious · 3 min walk to Hoegi Stn",
    "갬성있는방!!회기역8분": "Stylish room · 8 min to Hoegi Stn",
}


def _tr_station(s):
    for ko, en in _STATION_MAP:
        s = s.replace(ko, en)
    return s


def _tr_options(s):
    toks = [t.strip() for t in s.split(",")]
    return ", ".join(_OPTIONS_MAP.get(t, t) for t in toks if t)


def _tr_name(n):
    if n in _M2_TITLE_MAP:
        return _M2_TITLE_MAP[n]
    if ", " in n:
        parts = n.split(", ", 1)
        return f'{_TYPE_MAP.get(parts[0], parts[0])}, {_DONG_MAP.get(parts[1], parts[1])}'
    return n


def load_featured():
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, "listings.json")
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if not data.get("listings"):
            return [{**r, "total": ""} for r in FALLBACK_FEATURED]
        out = []
        for r in data["listings"]:
            out.append({
                "name": _tr_name(r["name"]),
                "platform": r["platform"],
                "type": _TYPE_MAP.get(r["type"], r["type"]),
                "rent": r["rent"],
                "total": r["total_display"],
                "deposit": r["deposit_display"],
                "size": r["size"],
                "station": _tr_station(r["station"]),
                "commute": _tr_station(r["commute"]),
                "options": _tr_options(r["options"]),
                "english": r["english"],
                "link": r["link"],
                "naver": r["naver"],
                "notes": r["notes"].replace("관리비", "maint. fee"),
            })
        return out
    except Exception:
        return [{**r, "total": ""} for r in FALLBACK_FEATURED]


# Neighborhood guide — stable, non-expiring search links.
areas = [
    ("Hoegi-dong", "회기동", "Hoegi Stn", "Line 1 · Gyeongui-Jungang · Gyeongchun",
     "5–15 min (15-min walk or bus Dongdaemun-01)", "₩450,000 – 700,000",
     "Hongneung Forest", "홍릉수목원", "경희대", "회기역",
     "Closest & most convenient. Best for direct Line-1 access."),
    ("Imun-dong / HUFS area", "이문동·외대앞", "Hankuk Univ. of FS Stn (Oedae-ap)", "Line 1",
     "~10–15 min (walk via KHU back gate)", "₩450,000 – 650,000",
     "Hankuk Univ. of Foreign Studies", "한국외국어대학교", "외대앞역", "외대앞역",
     "Back-gate access; lively student street, cheap eats."),
    ("Cheongnyangni", "청량리", "Cheongnyangni Stn", "Line 1 · Gyeongui-Jungang · Suin-Bundang · GTX-B(future)",
     "~15–20 min (1 stop + bus)", "₩500,000 – 750,000",
     "Seoul Yangnyeong Herbal Market / Gyeongdong Market", "서울약령시·경동시장", "청량리역", "경동시장",
     "Major transfer hub; markets, malls, Lotte Castle tower."),
    ("Sinimun", "신이문", "Sinimun Stn", "Line 1",
     "~20 min (1 stop to Hoegi + walk/bus)", "₩400,000 – 600,000",
     "Uireung Royal Tomb (UNESCO)", "의릉", "신이문역", "의릉",
     "Quiet & cheaper, still one stop from Hoegi."),
    ("Seokgye", "석계", "Seokgye Stn", "Line 1 · Line 6",
     "~25–30 min (2 stops on Line 1)", "₩450,000 – 650,000",
     "Gyeongchun Line Forest Trail", "경춘선숲길", "석계역", "경춘선숲길",
     "KEY Line-6 transfer point. Good if you want a Line-6 home + Line-1 commute."),
    ("Gwangwoon-dae / Wolgye", "광운대·월계", "Gwangwoon-dae Stn", "Line 1",
     "~25–30 min (3 stops on Line 1)", "₩450,000 – 650,000",
     "Gyeongchun Line Forest Trail", "경춘선숲길", "광운대", "경춘선숲길",
     "Budget pick on Line 1; calm residential."),
    ("Anam (Korea Univ.)", "안암", "Anam Stn", "Line 6",
     "~40–45 min (Line 6 → Seokgye transfer → Line 1)", "₩500,000 – 750,000",
     "Korea University", "고려대학교", "안암", "고려대학교",
     "Big student area on Line 6; lots of listings & food."),
    ("Bomun", "보문", "Bomun Stn", "Line 6 · Ui-Sinseol",
     "~45–50 min (Line 6 → Seokgye → Line 1)", "₩450,000 – 650,000",
     "Bomunsa Temple", "보문사", "보문역", "보문사",
     "Two lines; quiet, by Seongbukcheon stream."),
    ("Wolgok", "월곡", "Wolgok Stn", "Line 6",
     "~40–45 min (Line 6 → Seokgye → Line 1)", "₩400,000 – 600,000",
     "Dongduk Women's Univ.", "동덕여자대학교", "월곡역", "동덕여자대학교",
     "Cheapest Line-6 option; student neighborhood."),
]

# Platforms with English support.
platforms = [
    ("Ziptoss", "English-only site (/en)", "1 month+",
     "Negotiable", "Foreigners signing an English contract directly; biggest agency network",
     "https://ziptoss.com/en"),
    ("33m2 (삼삼엠투)", "English support, foreign card", "1 week+",
     "₩330,000 fixed", "Most listings; low fixed deposit; turnkey furnished short stays",
     "https://33m2.co.kr/landing/article_overseas"),
    ("LiveAnywhere", "EN / JP / CN", "1 month+",
     "Low / escrow", "Exchange & language students; instant foreign-card pay, email sign-up, escrow safety",
     "https://www.liveanywhere.me/en"),
    ("Wehome (위홈)", "English", "Short term+",
     "Varies", "Licensed legal homeshare; well-decorated whole units / studios",
     "https://www.wehome.me/"),
    ("Airbnb", "English", "Short term+",
     "Varies", "Fewer listings than before but still options; use the Monthly filter",
     "https://www.airbnb.com/s/Hoegi--Seoul/homes"),
]

suggestions = [
    ("Hoegi Station is your best bet", "It is a direct Line-1 stop AND served by Gyeongui-Jungang + Gyeongchun lines. KHU is a 15-min walk or one Dongdaemun-01 bus ride. Living near Hoegi keeps the commute under 20 minutes."),
    ("Keep the deposit low", "Without an ARC (Alien Registration Card, issued after arrival) landlords often want a bigger deposit. 33m2 (₩330k fixed), Ziptoss English contracts, and LiveAnywhere escrow are the safest low-deposit routes for foreigners."),
    ("One semester = watch the short-stay premium", "Turnkey short-stay platforms charge a monthly premium. If the ₩5M-deposit route is affordable, a normal studio via Ziptoss with a negotiated 4–6 month contract can be much cheaper per month."),
    ("Pay only through the platform / escrow", "Never wire money directly to a landlord before a signed contract. Check the property register (등기부등본) and use a licensed agent (공인중개사)."),
    ("Furnishing checklist", "Confirm: washer, fridge, A/C, induction/gas range, desk, bed, and WiFi are included before booking."),
    ("Hidden costs", "Ask whether maintenance fee (관리비) and utilities are separate. Korean winter gas heating can be expensive — confirm who pays."),
    ("Documents to bring", "Passport, financial proof, and (if available) ARC. Some short-stay platforms accept bookings without an ARC."),
]

# ---------------------------------------------------------------- CSS
DESKTOP_CSS = """\
:root{--bg:#f8f9fb;--panel:#ffffff;--soft:#f1f3f5;--line:#dee2e6;--text:#212529;--mut:#6c757d;--acc:#0055cc;--acc2:#1a7f4b;--warn:#664d03;--warn-bg:#fff8e1;--warn-border:#ffe082;--hero:#1a3a5c;--hero2:#0f2540}
*{box-sizing:border-box}
body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Noto Sans KR",sans-serif;background:var(--bg);color:var(--text);line-height:1.6;-webkit-text-size-adjust:100%}
.wrap{max-width:1280px;margin:0 auto;padding:24px 20px 72px}
header.hero{background:linear-gradient(135deg,var(--hero),var(--hero2));color:#fff;border-radius:18px;padding:28px 26px;margin-bottom:28px;box-shadow:0 2px 12px rgba(0,0,0,.12)}
header.hero h1{margin:0 0 6px;font-size:clamp(22px,4.5vw,32px);line-height:1.2;color:#fff}
header.hero p{margin:4px 0;color:rgba(255,255,255,.75);font-size:15px}
.badges{display:flex;flex-wrap:wrap;gap:8px;margin-top:16px}
.badge{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:#fff;padding:6px 14px;border-radius:999px;font-size:13px;font-weight:600}
.badge b{color:#7ee8a2}
section{margin:36px 0}
h2{font-size:clamp(18px,3.5vw,23px);margin:0 0 4px;color:var(--hero);display:flex;align-items:center;gap:9px}
h2 .dot{width:10px;height:10px;border-radius:50%;background:var(--acc)}
.sub{color:var(--mut);font-size:14px;margin:0 0 14px}
.table-outer{position:relative;border-radius:14px}
.tablewrap{overflow-x:auto;border:1px solid var(--line);border-radius:14px;background:var(--panel);box-shadow:0 1px 4px rgba(0,0,0,.06)}
.scroll-fade{position:absolute;top:0;right:0;bottom:0;width:90px;background:linear-gradient(to right,transparent,rgba(255,255,255,.97));border-radius:0 14px 14px 0;pointer-events:none;transition:opacity .25s;z-index:1}
.scroll-tip{position:absolute;top:10px;right:10px;background:var(--acc);color:#fff;font-size:11px;font-weight:700;padding:4px 10px;border-radius:999px;pointer-events:none;white-space:nowrap;transition:opacity .25s;z-index:2;box-shadow:0 1px 4px rgba(0,85,204,.3)}
.table-outer.at-end .scroll-fade{opacity:0}
.table-outer.at-end .scroll-tip{opacity:0}
table{width:100%;border-collapse:collapse;font-size:14px}
th,td{text-align:left;padding:12px 15px;border-bottom:1px solid var(--line);vertical-align:top}
th{background:var(--soft);color:var(--mut);font-weight:700;font-size:11.5px;text-transform:uppercase;letter-spacing:.5px;position:sticky;top:0;white-space:nowrap;user-select:none}
th.sortable{cursor:pointer}
th.sortable:hover{background:#e9ecef;color:var(--text)}
th.sort-asc::after{content:" ▲";font-size:10px}
th.sort-desc::after{content:" ▼";font-size:10px}
td.col-name{min-width:260px;font-weight:600;color:var(--hero);line-height:1.45}
td.col-name a:first-of-type{display:block;font-size:14px;color:var(--hero);margin-bottom:5px}
td.col-name a:first-of-type:hover{color:var(--acc)}
.row-num{display:inline-block;font-size:11px;font-weight:700;color:var(--mut);background:var(--soft);border:1px solid var(--line);border-radius:4px;padding:1px 6px;margin-bottom:4px;letter-spacing:.3px}
.row-links{display:flex;gap:8px;flex-wrap:wrap;margin-top:4px}
.map-btn{font-size:12px;padding:2px 8px;border-radius:6px;background:var(--soft);border:1px solid var(--line);color:var(--mut)!important;font-weight:600;white-space:nowrap}
.map-btn:hover{background:var(--line);color:var(--text)!important;text-decoration:none!important}
.plat-badge{font-size:11px;color:var(--mut);font-weight:500;display:block;margin:2px 0 3px}
.pg-bar{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:10px;font-size:13px;color:var(--mut)}
.pg-nav{display:flex;align-items:center;gap:6px}
.pg-btn{background:var(--soft);border:1px solid var(--line);color:var(--text);padding:4px 12px;border-radius:7px;cursor:pointer;font-size:13px;font-weight:600}
.pg-btn:disabled{opacity:.35;cursor:default}
.pg-btn:not(:disabled):hover{background:var(--line)}
#pg-pages{display:flex;gap:4px}
.pg-page{background:var(--soft);border:1px solid var(--line);color:var(--text);padding:4px 10px;border-radius:7px;cursor:pointer;font-size:13px;font-weight:600;min-width:34px;text-align:center}
.pg-page.active{background:var(--acc);color:#fff;border-color:var(--acc)}
.pg-page:not(.active):hover{background:var(--line)}
.pg-size-wrap select{margin-left:4px;padding:4px 8px;border:1px solid var(--line);border-radius:7px;font-size:13px;background:var(--soft);cursor:pointer}
.pg-info{color:var(--mut);font-size:13px}
th:nth-child(1){min-width:260px}
th:nth-child(2){min-width:110px}
th:nth-child(3),th:nth-child(4){min-width:130px}
th:nth-child(5){min-width:110px}
th:nth-child(6){min-width:70px}
th:nth-child(7){min-width:140px}
th:nth-child(8){min-width:130px}
th:nth-child(9){min-width:200px}
th:nth-child(10){min-width:180px}
tr:last-child td{border-bottom:none}
tr:hover td{background:#f8f9ff}
a{color:var(--acc);text-decoration:none;font-weight:600;white-space:nowrap}
a:hover{text-decoration:underline}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px}
.card{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:18px 20px;box-shadow:0 1px 3px rgba(0,0,0,.05)}
.card h3{margin:0 0 6px;font-size:15px;color:var(--acc2)}
.card p{margin:0;color:var(--mut);font-size:13.5px}
.note{background:var(--warn-bg);border:1px solid var(--warn-border);color:var(--warn);border-radius:12px;padding:12px 16px;font-size:13px;margin-top:12px}
footer{margin-top:48px;color:var(--mut);font-size:12.5px;border-top:1px solid var(--line);padding-top:18px}
@media (max-width:640px){
  .wrap{padding:12px 12px 48px}
  header.hero{border-radius:12px;padding:18px 16px}
  .badges{gap:6px;margin-top:12px}
  .badge{font-size:12px;padding:5px 10px}
  .note{font-size:12px;padding:10px 12px}
  h2{font-size:17px}
  .sub{font-size:13px}
  .pg-bar{flex-direction:column;align-items:flex-start;gap:8px}
  .pg-nav{flex-wrap:wrap}
  .pg-btn{min-height:40px;padding:6px 16px;font-size:14px}
  .pg-page{min-height:40px;min-width:40px;font-size:14px;padding:6px 8px}
  .pg-size-wrap{font-size:14px}
  .pg-size-wrap select{font-size:14px;padding:6px 10px;min-height:40px}
  #feat-table thead{display:none}
  #feat-table,#feat-table tbody,#feat-table tr,#feat-table td{display:block;width:100%}
  .table-outer:has(#feat-table) .tablewrap{overflow:visible;border:none;background:transparent;box-shadow:none}
  .table-outer:has(#feat-table) .scroll-fade,
  .table-outer:has(#feat-table) .scroll-tip{display:none}
  #feat-table tr{border:1px solid var(--line);border-radius:12px;margin:12px 0;background:var(--panel);box-shadow:0 1px 4px rgba(0,0,0,.07);overflow:hidden}
  #feat-table tr:hover td{background:transparent}
  #feat-table td{border:none;border-bottom:1px solid var(--line);padding:8px 14px;display:flex;justify-content:space-between;align-items:flex-start;gap:10px}
  #feat-table td:last-child{border-bottom:none}
  #feat-table td::before{content:attr(data-label);color:var(--mut);font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.3px;flex:0 0 36%;padding-top:2px}
  #feat-table td.col-name{flex-direction:column;align-items:flex-start;padding:12px 14px}
  #feat-table td.col-name::before{display:none}
  #feat-table td.col-name a:first-of-type{font-size:15px;font-weight:700;white-space:normal;word-break:break-word;line-height:1.4}
  #feat-table .plat-badge{font-size:12px}
  #feat-table .map-btn{font-size:13px;padding:6px 14px;min-height:36px;display:inline-flex;align-items:center}
  #feat-table .row-links{margin-top:8px}
  #feat-table td:nth-child(8),#feat-table td:nth-child(9){display:none}
  #feat-table td:nth-child(10){font-size:12.5px}
  .tablewrap{-webkit-overflow-scrolling:touch}
  .scroll-fade,.scroll-tip{display:none}
}"""

MOBILE_CSS = """\
:root{--bg:#f0f2f5;--panel:#ffffff;--soft:#e8eaed;--line:#d1d5db;--text:#1a1a2e;--mut:#6b7280;--acc:#1a56db;--acc2:#0e7c3f;--warn:#92400e;--warn-bg:#fffbeb;--warn-border:#fbbf24;--hero:#1a3a5c;--hero2:#0f2540}
*{box-sizing:border-box}
html{font-size:18px}
body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Noto Sans KR",sans-serif;background:var(--bg);color:var(--text);line-height:1.7;-webkit-text-size-adjust:100%}
.wrap{max-width:640px;margin:0 auto;padding:16px 14px 64px}
header.hero{background:linear-gradient(135deg,var(--hero),var(--hero2));color:#fff;border-radius:16px;padding:24px 20px;margin-bottom:24px;box-shadow:0 2px 12px rgba(0,0,0,.15)}
header.hero h1{margin:0 0 8px;font-size:26px;line-height:1.25;color:#fff}
header.hero p{margin:4px 0;color:rgba(255,255,255,.8);font-size:16px;line-height:1.6}
.badges{display:flex;flex-wrap:wrap;gap:8px;margin-top:14px}
.badge{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:#fff;padding:7px 14px;border-radius:999px;font-size:14px;font-weight:600}
.badge b{color:#7ee8a2}
.note{background:var(--warn-bg);border:1px solid var(--warn-border);color:var(--warn);border-radius:12px;padding:14px 16px;font-size:15px;margin-top:14px;line-height:1.6}
section{margin:28px 0}
h2{font-size:21px;margin:0 0 4px;color:var(--hero);display:flex;align-items:center;gap:9px}
h2 .dot{width:10px;height:10px;border-radius:50%;background:var(--acc);flex-shrink:0}
.sub{color:var(--mut);font-size:15px;margin:0 0 16px;line-height:1.6}
.table-outer{border-radius:0}
.tablewrap{overflow:visible;border:none;background:transparent;box-shadow:none}
.scroll-fade,.scroll-tip{display:none}
table{display:block;width:100%}
thead{display:none}
tbody{display:block;width:100%}
tr{display:block;border:1px solid var(--line);border-radius:14px;margin:14px 0;background:var(--panel);box-shadow:0 2px 6px rgba(0,0,0,.08);overflow:hidden}
tr:last-child td{border-bottom:none}
tr:hover td{background:transparent}
td{display:flex;width:100%;border:none;border-bottom:1px solid var(--line);padding:12px 16px;align-items:flex-start;gap:12px;min-height:48px;font-size:16px;vertical-align:top;box-sizing:border-box}
td:last-child{border-bottom:none}
td::before{content:attr(data-label);color:var(--mut);font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;flex:0 0 38%;padding-top:2px;line-height:1.5}
td.col-name{flex-direction:column;align-items:flex-start;padding:16px 16px 12px}
td.col-name::before{display:none}
td.col-name a:first-of-type{display:block;font-size:20px;font-weight:700;color:var(--hero);line-height:1.4;white-space:normal;word-break:break-word;margin-bottom:6px}
td.col-name a:first-of-type:hover{color:var(--acc)}
.row-num{display:inline-block;font-size:13px;font-weight:700;color:var(--mut);background:var(--soft);border:1px solid var(--line);border-radius:6px;padding:3px 10px;margin-bottom:8px}
.plat-badge{font-size:14px;color:var(--mut);font-weight:500;display:block;margin:2px 0 6px}
.row-links{display:flex;gap:10px;flex-wrap:wrap;margin-top:8px}
.map-btn{font-size:15px;padding:10px 18px;min-height:44px;border-radius:10px;background:var(--soft);border:1px solid var(--line);color:var(--mut)!important;font-weight:600;white-space:nowrap;display:inline-flex;align-items:center}
.map-btn:hover{background:var(--line);color:var(--text)!important;text-decoration:none!important}
a{color:var(--acc);text-decoration:none;font-weight:600}
a:hover{text-decoration:underline}
.pg-bar{display:flex;flex-direction:column;align-items:flex-start;gap:10px;margin-bottom:14px;font-size:16px;color:var(--mut)}
.pg-nav{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.pg-btn{background:var(--soft);border:1px solid var(--line);color:var(--text);padding:10px 20px;min-height:44px;border-radius:10px;cursor:pointer;font-size:16px;font-weight:600}
.pg-btn:disabled{opacity:.35;cursor:default}
.pg-btn:not(:disabled):hover{background:var(--line)}
#pg-pages{display:flex;gap:6px;flex-wrap:wrap}
.pg-page{background:var(--soft);border:1px solid var(--line);color:var(--text);padding:10px 14px;min-height:44px;min-width:44px;border-radius:10px;cursor:pointer;font-size:16px;font-weight:600;text-align:center;display:inline-flex;align-items:center;justify-content:center}
.pg-page.active{background:var(--acc);color:#fff;border-color:var(--acc)}
.pg-page:not(.active):hover{background:var(--line)}
.pg-size-wrap{font-size:16px}
.pg-size-wrap select{margin-left:6px;padding:8px 12px;border:1px solid var(--line);border-radius:10px;font-size:16px;background:var(--soft);cursor:pointer;min-height:44px}
.pg-info{font-size:15px;color:var(--mut)}
.cards{display:grid;grid-template-columns:1fr;gap:14px}
.card{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:20px 18px;box-shadow:0 1px 4px rgba(0,0,0,.06)}
.card h3{margin:0 0 8px;font-size:18px;color:var(--acc2);line-height:1.4}
.card p{margin:0;color:var(--mut);font-size:16px;line-height:1.7}
footer{margin-top:48px;color:var(--mut);font-size:14px;border-top:1px solid var(--line);padding-top:18px;line-height:1.7}
.desktop-link{display:inline-block;margin-top:14px;color:var(--acc);font-size:15px;font-weight:600}"""

# ---------------------------------------------------------------- HTML
def th(cells):
    return "<tr>" + "".join(f"<th>{c}</th>" for c in cells) + "</tr>"

def sortable_th(cells):
    return "<tr>" + "".join(f'<th class="sortable">{c}</th>' for c in cells) + "</tr>"

def td(label, val, is_link=False, link_text=None):
    if is_link and val:
        inner = f'<a href="{html.escape(val)}" target="_blank" rel="noopener">{html.escape(link_text or "Open")}</a>'
    else:
        inner = html.escape(str(val))
    return f'<td data-label="{html.escape(label)}">{inner}</td>'

def _build_body(mobile=False):
    featured = load_featured()
    F = []
    for i, r in enumerate(featured):
        name_cell = (
            f'<td data-label="Listing" class="col-name">'
            f'<span class="row-num">{i + 1}</span>'
            f'<a href="{html.escape(r["link"])}" target="_blank" rel="noopener">{html.escape(str(r["name"]))}</a>'
            f'<span class="plat-badge">{html.escape(str(r.get("platform", "")))}</span>'
            f'<span class="row-links">'
            f'<a href="{html.escape(r["naver"])}" target="_blank" rel="noopener" class="map-btn">🗺 Map</a>'
            f'</span></td>'
        )
        F.append("<tr>"
            + name_cell
            + td("Type", r["type"])
            + td("Monthly rent", r["rent"])
            + td("Total /mo", r.get("total", ""))
            + td("Deposit", r["deposit"])
            + td("Size", r["size"])
            + td("Station", r["station"])
            + td("Commute", r["commute"])
            + td("Options", r["options"])
            + td("Notes", r["notes"])
            + "</tr>")
    feat_rows = "\n".join(F)

    A = []
    for (en, ko, st, line, commute, rent, lm_en, lm_ko, zarea, narea, note) in areas:
        A.append("<tr>"
            + td("Area", f"{en} ({ko})")
            + td("Nearest station", f"{st}")
            + td("Line(s)", line)
            + td("Commute to KHU", commute)
            + td("Typical rent / mo", rent)
            + td("Nearest landmark", f"{lm_en} ({lm_ko})")
            + td("Ziptoss (EN)", ziptoss_en(zarea), True, "Search rooms")
            + td("Map", naver(narea), True, "Naver Map")
            + td("Landmark map", naver(lm_ko), True, "See landmark")
            + td("Notes", note)
            + "</tr>")
    area_rows = "\n".join(A)

    P = []
    for (nm, eng, mn, dep, best, url) in platforms:
        P.append("<tr>"
            + td("Platform", nm)
            + td("English", eng)
            + td("Min stay", mn)
            + td("Deposit", dep)
            + td("Best for", best)
            + td("Open", url, True, "Visit site")
            + "</tr>")
    plat_rows = "\n".join(P)

    sug = "\n".join(
        f'<div class="card"><h3>{html.escape(t)}</h3><p>{html.escape(d)}</p></div>'
        for t, d in suggestions)

    desktop_link = (
        '<br><a href="index.html?desktop=1" class="desktop-link">🖥 Desktop version &rarr;</a>'
        if mobile else ''
    )

    return f"""
<div class="wrap">
<header class="hero">
  <h1>🏠 Rooms near Kyung Hee University</h1>
  <p>A housing guide for a one-semester exchange student — Seoul Campus (Hoegi-dong, Dongdaemun-gu).</p>
  <div class="badges">
    <span class="badge">🚇 Subway <b>Line 1 / Line 6</b></span>
    <span class="badge">⏱ Commute <b>≤ 1 hour</b></span>
    <span class="badge">📅 Stay <b>4–6 months</b></span>
    <span class="badge">💰 Maint. fee incl. <b>~₩1,200,000 / mo</b></span>
    <span class="badge">🗣 <b>English-friendly</b> platforms</span>
    <span class="badge">🎯 Priority: <b>lower price</b></span>
  </div>
  <div class="note">⚠ Prices are a snapshot from {SNAPSHOT}. Individual listing links can expire within days; the neighborhood "Search rooms" links stay live. Always verify the current price on the platform.</div>
</header>

<section>
  <h2><span class="dot"></span>Featured live listings <span style="font-weight:400;color:var(--mut);font-size:13px">(snapshot {SNAPSHOT})</span></h2>
  <p class="sub">Live units from 33m2 + Ziptoss + Airbnb near KHU, sorted by total monthly cost (rent + maint. fee) ≤ ₩1,200,000. Open the link to confirm availability.</p>
  <div class="pg-bar">
    <span class="pg-info" id="pg-info"></span>
    <span class="pg-nav">
      <button class="pg-btn" id="pg-prev">&#8249; Prev</button>
      <span id="pg-pages"></span>
      <button class="pg-btn" id="pg-next">Next &#8250;</button>
    </span>
    <span class="pg-size-wrap">Show:
      <select id="pg-size">
        <option value="10">10</option>
        <option value="20">20</option>
        <option value="30" selected>30</option>
        <option value="50">50</option>
        <option value="100">100</option>
        <option value="0">All</option>
      </select>
    </span>
  </div>
  <div class="table-outer"><div class="scroll-fade"></div><div class="scroll-tip">scroll ›</div><div class="tablewrap"><table id="feat-table">
  <thead>{sortable_th(["Listing + Links","Type","Monthly rent","Total /mo","Deposit","Size","Station","Commute","Options","Notes"])}</thead>
  <tbody>{feat_rows}</tbody>
  </table></div></div>
</section>

<section>
  <h2><span class="dot"></span>Neighborhood guide — live search links</h2>
  <p class="sub">Each row links to live, filtered listings (Ziptoss English) plus a Naver Map pin for the area and its landmark. Sorted closest → cheaper. All are within 1 hour of KHU.</p>
  <div class="table-outer"><div class="scroll-fade"></div><div class="scroll-tip">scroll ›</div><div class="tablewrap"><table>
  {th(["Area","Nearest station","Line(s)","Commute to KHU","Typical rent / mo","Nearest landmark","Ziptoss (EN)","Map","Landmark map","Notes"])}
  {area_rows}
  </table></div></div>
</section>

<section>
  <h2><span class="dot"></span>English-friendly platforms</h2>
  <p class="sub">Recommended order for an exchange student. Ziptoss & 33m2 first; LiveAnywhere for the safest foreigner flow.</p>
  <div class="table-outer"><div class="scroll-fade"></div><div class="scroll-tip">scroll ›</div><div class="tablewrap"><table>
  {th(["Platform","English","Min stay","Deposit","Best for","Open"])}
  {plat_rows}
  </table></div></div>
</section>

<section>
  <h2><span class="dot"></span>My extra suggestions</h2>
  <p class="sub">Practical tips to save money and avoid trouble.</p>
  <div class="cards">
  {sug}
  </div>
</section>

<footer>
  Built {SNAPSHOT}. Rent figures are approximate, deposit-based studio ranges; short-stay furnished units (33m2 / LiveAnywhere, ₩330k deposit) run higher per month but need no large deposit.
  Live listings scraped from: 33m2, Ziptoss, Airbnb. Airbnb prices are per-night×30 estimates only. Commute times estimated from Hoegi Station.{desktop_link}
</footer>
</div>
<script>
(function(){{
  /* ---- scroll fade indicators ---- */
  document.querySelectorAll('.table-outer').forEach(function(outer) {{
    var wrap = outer.querySelector('.tablewrap');
    if (!wrap) return;
    function upd() {{
      var atEnd = wrap.scrollLeft + wrap.clientWidth >= wrap.scrollWidth - 4;
      outer.classList.toggle('at-end', atEnd);
    }}
    wrap.addEventListener('scroll', upd, {{passive:true}});
    window.addEventListener('resize', upd, {{passive:true}});
    upd();
  }});

  /* ---- sortable + paginated featured table ---- */
  var tbl = document.getElementById('feat-table');
  if (!tbl) return;
  var thead = tbl.querySelector('thead tr');
  var tbody = tbl.querySelector('tbody');
  var allRows = Array.from(tbody.rows);
  var sortCol = -1, sortAsc = true;
  var curPage = 1;
  var pageSize = 30;

  function cellVal(row, col) {{
    var c = row.cells[col];
    return c ? c.textContent.trim() : '';
  }}
  function numOrStr(v) {{
    var n = parseFloat(v.replace(/[₩, ]/g,'').split('/')[0]);
    return isNaN(n) ? v : n;
  }}

  function renderPage() {{
    var total = allRows.length;
    var size = (pageSize === 0) ? total : pageSize;
    var totalPages = (pageSize === 0) ? 1 : Math.ceil(total / size);
    if (curPage > totalPages) curPage = totalPages;
    if (curPage < 1) curPage = 1;
    var start = (curPage - 1) * size;
    var end = (pageSize === 0) ? total : Math.min(start + size, total);

    allRows.forEach(function(r) {{ r.style.display = 'none'; }});
    allRows.slice(start, end).forEach(function(r) {{ r.style.display = ''; }});

    var info = document.getElementById('pg-info');
    if (info) info.textContent = (start+1) + '–' + end + ' / ' + total + ' listings';

    var pagesEl = document.getElementById('pg-pages');
    if (pagesEl) {{
      pagesEl.innerHTML = '';
      if (totalPages > 1) {{
        var range = [];
        var delta = 2;
        for (var i = Math.max(1, curPage-delta); i <= Math.min(totalPages, curPage+delta); i++) range.push(i);
        if (range[0] > 1) {{ addPageBtn(pagesEl, 1); if (range[0] > 2) addDot(pagesEl); }}
        range.forEach(function(p) {{ addPageBtn(pagesEl, p); }});
        if (range[range.length-1] < totalPages) {{
          if (range[range.length-1] < totalPages-1) addDot(pagesEl);
          addPageBtn(pagesEl, totalPages);
        }}
      }}
    }}

    var prev = document.getElementById('pg-prev');
    var next = document.getElementById('pg-next');
    if (prev) prev.disabled = (curPage <= 1);
    if (next) next.disabled = (curPage >= totalPages || totalPages <= 1);
  }}

  function addPageBtn(container, p) {{
    var btn = document.createElement('button');
    btn.className = 'pg-page' + (p === curPage ? ' active' : '');
    btn.textContent = p;
    btn.addEventListener('click', function() {{ curPage = p; renderPage(); }});
    container.appendChild(btn);
  }}
  function addDot(container) {{
    var s = document.createElement('span');
    s.textContent = '…';
    s.style.cssText = 'padding:4px 2px;color:var(--mut);font-size:13px';
    container.appendChild(s);
  }}

  document.getElementById('pg-prev') && document.getElementById('pg-prev').addEventListener('click', function() {{ curPage--; renderPage(); }});
  document.getElementById('pg-next') && document.getElementById('pg-next').addEventListener('click', function() {{ curPage++; renderPage(); }});
  var sizeEl = document.getElementById('pg-size');
  if (sizeEl) {{
    sizeEl.addEventListener('change', function() {{
      pageSize = parseInt(this.value);
      curPage = 1;
      renderPage();
    }});
  }}

  if (thead) {{
    Array.from(thead.cells).forEach(function(th, i) {{
      th.addEventListener('click', function() {{
        var asc = (sortCol === i) ? !sortAsc : true;
        allRows.sort(function(a, b) {{
          var av = numOrStr(cellVal(a, i));
          var bv = numOrStr(cellVal(b, i));
          if (typeof av === 'number' && typeof bv === 'number') return asc ? av-bv : bv-av;
          return asc ? String(av).localeCompare(String(bv),'ko') : String(bv).localeCompare(String(av),'ko');
        }});
        allRows.forEach(function(r) {{ tbody.appendChild(r); }});
        Array.from(thead.cells).forEach(function(h) {{
          h.classList.remove('sort-asc','sort-desc');
        }});
        th.classList.add(asc ? 'sort-asc' : 'sort-desc');
        sortCol = i; sortAsc = asc;
        curPage = 1;
        renderPage();
      }});
    }});
  }}

  renderPage();
}})();
</script>
"""


def build_desktop_html():
    body = _build_body(mobile=False)
    return (
        "<!DOCTYPE html>\n"
        '<html lang="en">\n'
        "<head>\n"
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">\n'
        "<title>Rooms near Kyung Hee University — Exchange Student Guide</title>\n"
        "<script>if(window.matchMedia&&matchMedia('(max-width:640px)').matches"
        "&&location.search.indexOf('desktop')<0){location.replace('mobile.html');}</script>\n"
        "<style>\n" + DESKTOP_CSS + "\n</style>\n"
        "</head>\n"
        "<body>\n"
        + body +
        "\n</body>\n</html>"
    )


def build_mobile_html():
    body = _build_body(mobile=True)
    return (
        "<!DOCTYPE html>\n"
        '<html lang="en">\n'
        "<head>\n"
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">\n'
        "<title>Rooms near KHU — Mobile Guide</title>\n"
        "<script>if(window.matchMedia&&matchMedia('(min-width:641px)').matches)"
        "{location.replace('index.html');}</script>\n"
        "<style>\n" + MOBILE_CSS + "\n</style>\n"
        "</head>\n"
        "<body>\n"
        + body +
        "\n</body>\n</html>"
    )


# ---------------------------------------------------------------- EXCEL
def build_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()

    head_fill = PatternFill("solid", fgColor="16324A")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="16324A")
    link_font = Font(color="0563C1", underline="single")
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws, headers, rows, link_cols, title):
        ws.cell(1, 1, title).font = title_font
        ws.cell(2, 1, f"Snapshot {SNAPSHOT} — prices change; verify on platform.").font = Font(italic=True, color="888888")
        hr = 4
        for c, h in enumerate(headers, 1):
            cell = ws.cell(hr, c, h)
            cell.fill = head_fill; cell.font = head_font; cell.alignment = wrap; cell.border = border
        for ri, row in enumerate(rows, hr + 1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.alignment = wrap; cell.border = border
                if ci in link_cols and isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val; cell.value = "Open link"; cell.font = link_font
        ws.freeze_panes = ws.cell(hr + 1, 1)

    # Sheet 1: Featured
    ws1 = wb.active; ws1.title = "Featured listings"
    featured = load_featured()
    h1 = ["Listing","Type","Monthly rent","Total /mo","Deposit","Size","Station",
          "Commute","Options","Listing link","Naver Map","Notes"]
    r1 = [[r["name"],r["type"],r["rent"],r.get("total",""),r["deposit"],r["size"],r["station"],
           r["commute"],r["options"],r["link"],r["naver"],r["notes"]] for r in featured]
    style_sheet(ws1, h1, r1, {10,11}, "Featured live listings — near Kyung Hee University")
    widths1 = [36,16,22,18,16,10,28,26,50,14,14,42]
    for i,w in enumerate(widths1,1): ws1.column_dimensions[chr(64+i) if i<=26 else 'A'].width = w

    # Sheet 2: Neighborhoods
    ws2 = wb.create_sheet("Neighborhood guide")
    h2 = ["Area","Nearest station","Line(s)","Commute to KHU","Typical rent / mo",
          "Nearest landmark","Ziptoss (EN) search","Area map","Landmark map","Notes"]
    r2 = [[f"{en} ({ko})", st, line, commute, rent, f"{lm_en} ({lm_ko})",
           ziptoss_en(zarea), naver(narea), naver(lm_ko), note]
          for (en,ko,st,line,commute,rent,lm_en,lm_ko,zarea,narea,note) in areas]
    style_sheet(ws2, h2, r2, {7,8,9}, "Neighborhood guide — live search links")
    widths2 = [24,30,30,34,20,34,20,14,14,40]
    for i,w in enumerate(widths2,1): ws2.column_dimensions[chr(64+i)].width = w

    # Sheet 3: Platforms
    ws3 = wb.create_sheet("Platforms")
    h3 = ["Platform","English","Min stay","Deposit","Best for","Open"]
    r3 = [[nm,eng,mn,dep,best,url] for (nm,eng,mn,dep,best,url) in platforms]
    style_sheet(ws3, h3, r3, {6}, "English-friendly rental platforms")
    widths3 = [18,26,12,16,52,12]
    for i,w in enumerate(widths3,1): ws3.column_dimensions[chr(64+i)].width = w

    # Sheet 4: Suggestions
    ws4 = wb.create_sheet("Tips")
    style_sheet(ws4, ["Tip","Detail"], [[t,d] for t,d in suggestions], set(), "Extra suggestions")
    ws4.column_dimensions['A'].width = 36; ws4.column_dimensions['B'].width = 90

    wb.save(path)

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(here, "index.html"), "w", encoding="utf-8") as f:
        f.write(build_desktop_html())
    with open(os.path.join(here, "mobile.html"), "w", encoding="utf-8") as f:
        f.write(build_mobile_html())
    build_xlsx(os.path.join(here, "rooms.xlsx"))
    print("OK -> index.html, mobile.html, rooms.xlsx")
